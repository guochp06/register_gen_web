"""
层级化寄存器解析器 - 支持addr_map和register两种表格类型
完整移植自 gen_rdl_from_xls.py
"""
import xlrd
import openpyxl
import re
import os
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from pathlib import Path

from app.services.excel_validator import ExcelValidator


@dataclass
class RegisterField:
    name: str
    msb: int
    lsb: int
    access: str
    reset_value: str
    description: str = ""


@dataclass
class Register:
    name: str
    offset: int
    width: int
    fields: List[RegisterField] = field(default_factory=list)
    description: str = ""
    is_array: bool = False
    array_count: int = 1


@dataclass
class Module:
    name: str
    start_addr: int
    end_addr: int
    size: int
    registers: List[Register] = field(default_factory=list)
    submodules: List['Module'] = field(default_factory=list)
    is_array: bool = False
    array_count: int = 1
    source_file: str = ""
    description: str = ""
    is_array_instance: bool = False  # 是否是模块数组的实例
    base_module_name: str = ""  # 对应的基础模块名称（用于数组实例）


@dataclass
class RegisterHierarchy:
    """完整的寄存器层级结构"""
    version_name: str  # 用户输入的版本名（用于数据库和文件路径）
    top_addrmap_name: str = ""  # 顶层addr_map名字（从Excel解析，如soc_addr_map）
    top_modules: List[Module] = field(default_factory=list)
    all_modules: Dict[str, Module] = field(default_factory=dict)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


class HierarchyParser:
    """层级化寄存器解析器 - 完整实现"""

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.all_modules = {}
        self.addr_map_tables = {}
        self.register_tables = {}
        self.validator = ExcelValidator()

    def parse_files(self, file_paths: List[str], version_name: str) -> RegisterHierarchy:
        """解析一组Excel文件，构建完整层级结构"""
        self.errors = []
        self.warnings = []
        self.all_modules = {}
        self.addr_map_tables = {}
        self.register_tables = {}
        self.validator.clear()

        hierarchy = RegisterHierarchy(version_name=version_name)

        # 第一步：扫描所有文件，分类表格
        for file_path in file_paths:
            self._scan_file(file_path)

        # 第二步：找到顶层addr_map
        top_addr_map = self._find_top_addr_map()
        if not top_addr_map:
            # 没有顶层addr_map，所有register表作为独立模块（作为顶层）
            # 使用第一个register表名作为top_addrmap_name
            if self.register_tables:
                first_reg_name = list(self.register_tables.keys())[0]
                hierarchy.top_addrmap_name = first_reg_name

            for name, (file_path, sheet_name, data) in self.register_tables.items():
                # 如果已有错误，停止处理后续模块
                if self.errors:
                    break

                registers = self._parse_register_sheet(data, name, 0)
                # 如果有错误，停止处理
                if self.errors:
                    break
                if registers:
                    module = Module(
                        name=name,
                        start_addr=0,
                        end_addr=0,
                        size=0,
                        registers=registers,
                        source_file=file_path
                    )
                    hierarchy.top_modules.append(module)
                    hierarchy.all_modules[name] = module
                    self.all_modules[name] = module  # Also update self.all_modules
        else:
            # 从顶层addr_map构建层级
            # 创建顶层soc_addr_map模块，包含所有子模块
            top_module = self._build_top_addrmap_module(top_addr_map)
            if top_module:
                hierarchy.top_modules = [top_module]
                self.all_modules[top_module.name] = top_module
            else:
                hierarchy.top_modules = []
            # 使用顶层addr_map的sheet名作为顶层addrmap名
            _, top_sheet_name, _ = top_addr_map
            hierarchy.top_addrmap_name = top_sheet_name

        # 合并验证器的错误和警告
        val_errors, val_warnings = self.validator.get_results()
        self.errors.extend(val_errors)
        self.warnings.extend(val_warnings)

        hierarchy.errors = self.errors
        hierarchy.warnings = self.warnings
        hierarchy.all_modules = self.all_modules

        return hierarchy

    def _scan_file(self, file_path: str):
        """扫描单个Excel文件，分类表格"""
        try:
            path = Path(file_path)
            if not path.exists():
                self.errors.append(f"文件不存在: {file_path}")
                return

            try:
                workbook = xlrd.open_workbook(file_path)
                self._scan_workbook_xlrd(workbook, file_path)
            except Exception:
                try:
                    workbook = openpyxl.load_workbook(file_path)
                    self._scan_workbook_openpyxl(workbook, file_path)
                except Exception as e:
                    self.errors.append(f"无法解析文件 {file_path}: {str(e)}")
        except Exception as e:
            self.errors.append(f"扫描文件出错 {file_path}: {str(e)}")

    def _scan_workbook_xlrd(self, workbook: xlrd.book.Book, file_path: str):
        """扫描xlrd工作簿"""
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_name = sheet.name

            if sheet_name.lower() in ['note', 'notes', '说明', '注释']:
                continue

            headers = self._detect_headers_xlrd(sheet)
            if not headers:
                continue

            table_type = self._detect_table_type(headers)

            if table_type == 'addr_map':
                key = sheet_name
                self.addr_map_tables[key] = (file_path, sheet_name, {'xlrd': sheet, 'headers': headers})
            elif table_type == 'register':
                key = sheet_name
                self.register_tables[key] = (file_path, sheet_name, {'xlrd': sheet, 'headers': headers})

    def _scan_workbook_openpyxl(self, workbook: openpyxl.Workbook, file_path: str):
        """扫描openpyxl工作簿"""
        for sheet in workbook.worksheets:
            sheet_name = sheet.title

            if sheet_name.lower() in ['note', 'notes', '说明', '注释']:
                continue

            headers = self._detect_headers_openpyxl(sheet)
            if not headers:
                continue

            table_type = self._detect_table_type(headers)

            if table_type == 'addr_map':
                key = sheet_name
                self.addr_map_tables[key] = (file_path, sheet_name, {'openpyxl': sheet, 'headers': headers})
            elif table_type == 'register':
                key = sheet_name
                self.register_tables[key] = (file_path, sheet_name, {'openpyxl': sheet, 'headers': headers})

    def _detect_headers_xlrd(self, sheet: xlrd.sheet.Sheet) -> Dict[str, int]:
        """检测xlrd表格列头"""
        headers = {}
        expected = {
            'module_name': ['module_name', 'module name', 'modulename', '模块名', 'module', 'Module', 'MODULE_NAME'],
            'start_addr': ['start_addr', 'start addr', 'startaddr', 'startaddress', 'base_addr', 'base addr', 'baseaddr', '起始地址', 'start', 'Start', 'START_ADDR', 'Start Address', 'BASE_ADDR'],
            'end_addr': ['end_addr', 'end addr', 'endaddr', 'endaddress', '结束地址', 'end', 'End', 'END_ADDR', 'End Address'],
            'size': ['size', '大小', 'Size', 'SIZE'],
            'offset_address': ['offsetaddress', 'offset address', 'offset', '地址偏移', 'offset_address', 'OffsetAddress', 'Offset'],
            'reg_name': ['regname', 'reg name', '寄存器名', 'register', 'Register', 'RegName', 'REG_NAME', 'Register Name', 'registername'],
            'width': ['width', '位宽', 'Width', 'WIDTH'],
            'bits': ['bits', 'bit range', '位范围', 'bit_range', 'BitRange', 'Bits', 'BIT', 'field_pos', 'fieldpos', 'FIELD_POS', 'FieldPos'],
            'msb': ['msb', 'MSB', '最高位'],
            'lsb': ['lsb', 'LSB', '最低位'],
            'field_name': ['fieldname', 'field name', '字段名', 'field', 'Field', 'FieldName', 'FIELD_NAME', 'Field Name'],
            'access': ['access', '访问类型', 'accesstype', 'Access', 'ACCESS'],
            'reset_value': ['resetvalue', 'reset value', '复位值', 'reset', 'ResetValue', 'Reset', 'Reset Value'],
            'description': ['description', 'desc', '描述', 'Description', 'DESCRIPTION']
        }

        for col_idx in range(sheet.ncols):
            cell_value = str(sheet.cell_value(0, col_idx)).strip()
            cell_lower = cell_value.lower().replace(' ', '').replace('_', '')
            for key, aliases in expected.items():
                aliases_norm = [a.lower().replace(' ', '').replace('_', '') for a in aliases]
                if cell_lower in aliases_norm:
                    headers[key] = col_idx

        return headers

    def _detect_headers_openpyxl(self, sheet) -> Dict[str, int]:
        """检测openpyxl表格列头"""
        headers = {}
        expected = {
            'module_name': ['module_name', 'module name', 'modulename', '模块名', 'module', 'Module', 'MODULE_NAME'],
            'start_addr': ['start_addr', 'start addr', 'startaddr', 'startaddress', 'base_addr', 'base addr', 'baseaddr', '起始地址', 'start', 'Start', 'START_ADDR', 'Start Address', 'BASE_ADDR'],
            'end_addr': ['end_addr', 'end addr', 'endaddr', 'endaddress', '结束地址', 'end', 'End', 'END_ADDR', 'End Address'],
            'size': ['size', '大小', 'Size', 'SIZE'],
            'offset_address': ['offsetaddress', 'offset address', 'offset', '地址偏移', 'offset_address', 'OffsetAddress', 'Offset'],
            'reg_name': ['regname', 'reg name', '寄存器名', 'register', 'Register', 'RegName', 'REG_NAME', 'Register Name', 'registername'],
            'width': ['width', '位宽', 'Width', 'WIDTH'],
            'bits': ['bits', 'bit range', '位范围', 'bit_range', 'BitRange', 'Bits', 'BIT', 'field_pos', 'fieldpos', 'FIELD_POS', 'FieldPos'],
            'msb': ['msb', 'MSB', '最高位'],
            'lsb': ['lsb', 'LSB', '最低位'],
            'field_name': ['fieldname', 'field name', '字段名', 'field', 'Field', 'FieldName', 'FIELD_NAME', 'Field Name'],
            'access': ['access', '访问类型', 'accesstype', 'Access', 'ACCESS'],
            'reset_value': ['resetvalue', 'reset value', '复位值', 'reset', 'ResetValue', 'Reset', 'Reset Value'],
            'description': ['description', 'desc', '描述', 'Description', 'DESCRIPTION']
        }

        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        for col_idx, cell_value in enumerate(header_row):
            if cell_value:
                cell_str = str(cell_value).strip()
                cell_lower = cell_str.lower().replace(' ', '').replace('_', '')
                for key, aliases in expected.items():
                    aliases_norm = [a.lower().replace(' ', '').replace('_', '') for a in aliases]
                    if cell_lower in aliases_norm:
                        headers[key] = col_idx

        return headers

    def _detect_table_type(self, headers: Dict[str, int]) -> str:
        """检测表格类型"""
        if 'module_name' in headers and 'start_addr' in headers:
            return 'addr_map'
        elif 'offset_address' in headers and 'reg_name' in headers:
            return 'register'
        return 'unknown'

    def _find_top_addr_map(self) -> Optional[Tuple[str, str, Dict]]:
        """找到顶层addr_map - 优先选择引用最多其他addr_map的那个"""
        # 找到所有被引用的模块名（包括数组实例的基模块名）
        referenced_modules = set()
        addr_map_refs = {}  # addr_map_name -> list of modules it references

        for key, (file_path, sheet_name, data) in self.addr_map_tables.items():
            modules = self._parse_addr_map_modules(data)
            ref_list = []
            for mod_name, _, _ in modules:
                # 获取基模块名（处理数组实例如 PEC0 -> PEC）
                _, _, base_name = self._parse_module_array(mod_name)
                referenced_modules.add(base_name)
                ref_list.append(base_name)
            addr_map_refs[key] = ref_list

        # 找不被引用的addr_map作为候选顶层
        candidates = []
        for key, (file_path, sheet_name, data) in self.addr_map_tables.items():
            if key not in referenced_modules:
                candidates.append((key, file_path, sheet_name, data, len(addr_map_refs.get(key, []))))

        if not candidates:
            # 如果没有不被引用的，返回第一个
            if self.addr_map_tables:
                first = list(self.addr_map_tables.values())[0]
                return first
            return None

        if len(candidates) == 1:
            key, file_path, sheet_name, data, _ = candidates[0]
            return (file_path, sheet_name, data)

        # 有多个候选，选择引用最多的那个
        # 按引用数量降序排序
        candidates.sort(key=lambda x: x[4], reverse=True)

        best_candidate = candidates[0]
        best_key, best_file, best_sheet, best_data, best_count = best_candidate

        # 如果有多个候选且最佳候选的引用数明显多于其他，发出警告
        if len(candidates) > 1:
            second_best = candidates[1]
            second_key, _, _, _, second_count = second_best

            if best_count > second_count:
                self.warnings.append(
                    f"发现多个候选顶层addr_map，选择引用最多的 '{best_key}' "
                    f"(引用{best_count}个模块) 作为顶层，"
                    f"而非 '{second_key}' (引用{second_count}个模块)。"
                    f"层次结构已自动调整为以 '{best_key}' 为根。"
                )
            else:
                # 引用数相同，选择第一个但警告用户
                other_candidates = [c[0] for c in candidates[1:]]
                self.warnings.append(
                    f"多个addr_map引用数相同({best_count}个)，选择 '{best_key}' 作为顶层。"
                    f"其他候选: {', '.join(other_candidates)}。"
                    f"如需使用其他顶层，请单独上传该文件。"
                )

        return (best_file, best_sheet, best_data)

    def _parse_addr_map_modules(self, data: Dict) -> List[Tuple[str, int, int]]:
        """解析addr_map中的模块列表"""
        modules = []
        sheet = data.get('xlrd') or data.get('openpyxl')
        headers = data['headers']

        if 'xlrd' in data:
            for row_idx in range(1, sheet.nrows):
                mod_name = str(sheet.cell_value(row_idx, headers.get('module_name', 0))).strip()
                start_str = str(sheet.cell_value(row_idx, headers.get('start_addr', 1))).strip()

                if not mod_name:
                    continue

                start_addr = self._parse_address(start_str)
                modules.append((mod_name, start_addr, row_idx))
        else:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                mod_name = str(row[headers.get('module_name', 0)]).strip() if row[headers.get('module_name', 0)] else ""
                start_str = str(row[headers.get('start_addr', 1)]).strip() if row[headers.get('start_addr', 1)] else ""

                if not mod_name:
                    continue

                start_addr = self._parse_address(start_str)
                modules.append((mod_name, start_addr, row_idx))

        return modules

    def _build_top_addrmap_module(self, top_addr_map: Tuple[str, str, Dict]) -> Optional[Module]:
        """构建顶层addr_map模块（如soc_addr_map），包含所有子模块"""
        file_path, sheet_name, data = top_addr_map

        sheet = data.get('xlrd') or data.get('openpyxl')
        headers = data['headers']

        # 计算顶层模块的地址范围
        min_addr = float('inf')
        max_addr = 0

        # 首先收集所有子模块信息
        child_modules = []

        rows = []
        if 'xlrd' in data:
            for row_idx in range(1, sheet.nrows):
                rows.append((row_idx, None))
        else:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                rows.append((row_idx, row))

        for row_idx, row_data in rows:
            if self.errors:
                return None

            # 获取模块名称和地址
            if 'xlrd' in data:
                mod_name = str(sheet.cell_value(row_idx, headers.get('module_name', 0))).replace(' ', '').replace('（', '(').replace('）', ')')
                start_str = str(sheet.cell_value(row_idx, headers.get('start_addr', 1))).replace(' ', '')
                end_str = str(sheet.cell_value(row_idx, headers.get('end_addr', 2))).replace(' ', '') if 'end_addr' in headers else ""
            else:
                mod_name = str(row_data[headers.get('module_name', 0)]).replace(' ', '').replace('（', '(').replace('）', ')') if row_data[headers.get('module_name', 0)] else ""
                start_str = str(row_data[headers.get('start_addr', 1)]).replace(' ', '') if row_data[headers.get('start_addr', 1)] else ""
                end_str = str(row_data[headers.get('end_addr', 2)]).replace(' ', '') if 'end_addr' in headers and row_data[headers.get('end_addr', 2)] else ""

            if not mod_name:
                continue

            start_addr = self._parse_address(start_str)
            end_addr = self._parse_address(end_str) if end_str else 0

            if start_addr < min_addr:
                min_addr = start_addr
            if end_addr > max_addr:
                max_addr = end_addr

            # 解析模块并添加到子模块列表
            if 'xlrd' in data:
                row_modules = self._parse_addr_map_row_xlrd_expanded(sheet, headers, row_idx, file_path, 0, sheet_name)
            else:
                row_modules = self._parse_addr_map_row_openpyxl_expanded(row_data, headers, row_idx, file_path, 0, sheet_name)

            if row_modules:
                child_modules.extend(row_modules)

        # After collecting all child modules, propagate registers/submodules from base to instances
        for i, module in enumerate(child_modules):
            base_name = getattr(module, 'base_module_name', None)
            if base_name and base_name in self.all_modules:
                base_module = self.all_modules[base_name]
                # Copy registers/submodules from base if instance doesn't have them
                needs_regs = base_module.registers and not module.registers
                needs_submods = base_module.submodules and not module.submodules
                if needs_regs or needs_submods:
                    child_modules[i] = Module(
                        name=module.name,
                        start_addr=module.start_addr,
                        end_addr=module.end_addr,
                        size=module.size,
                        registers=base_module.registers if needs_regs else module.registers,
                        submodules=base_module.submodules if needs_submods else module.submodules,
                        is_array=module.is_array,
                        array_count=module.array_count,
                        source_file=module.source_file,
                        description=module.description,
                        is_array_instance=module.is_array_instance,
                        base_module_name=module.base_module_name
                    )
                    # Update all_modules
                    self.all_modules[module.name] = child_modules[i]

        # 创建顶层模块
        if min_addr == float('inf'):
            min_addr = 0

        # Calculate size from submodules (includes RALF modules with proper sizes)
        if child_modules:
            submodule_min = min(m.start_addr for m in child_modules)
            submodule_max = max(m.end_addr for m in child_modules)
            min_addr = min(min_addr, submodule_min)
            max_addr = max(max_addr, submodule_max)

        top_module = Module(
            name=sheet_name,
            start_addr=min_addr,
            end_addr=max_addr,
            size=max_addr - min_addr + 1 if max_addr >= min_addr else 0,
            registers=[],  # addr_map没有寄存器，只有子模块
            submodules=child_modules,
            source_file=file_path
        )

        return top_module

    def _build_hierarchy_from_addr_map(self, top_addr_map: Tuple[str, str, Dict], parent_base_addr: int = 0, parent_size: int = 0) -> List[Module]:
        """从顶层addr_map构建层级结构，支持模块数组展开"""
        file_path, sheet_name, data = top_addr_map
        modules = []

        sheet = data.get('xlrd') or data.get('openpyxl')
        headers = data['headers']

        # Calculate parent module size if not provided (for submodules)
        if not parent_size and 'end_addr' in headers:
            # Find the maximum end_addr to determine parent size
            max_end = 0
            if 'xlrd' in data:
                for row_idx in range(1, sheet.nrows):
                    end_str = str(sheet.cell_value(row_idx, headers.get('end_addr', 2))).replace(' ', '')
                    if end_str:
                        try:
                            end = self._parse_address(end_str)
                            max_end = max(max_end, end)
                        except:
                            pass
            else:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    end_str = str(row[headers.get('end_addr', 2)]).replace(' ', '') if row[headers.get('end_addr', 2)] else ""
                    if end_str:
                        try:
                            end = self._parse_address(end_str)
                            max_end = max(max_end, end)
                        except:
                            pass
            parent_size = max_end + 1 if max_end > 0 else 0

        rows = []
        if 'xlrd' in data:
            for row_idx in range(1, sheet.nrows):
                rows.append((row_idx, None))
        else:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                rows.append((row_idx, row))

        for row_idx, row_data in rows:
            # 如果有错误，停止处理
            if self.errors:
                return []

            if 'xlrd' in data:
                row_modules = self._parse_addr_map_row_xlrd_expanded(sheet, headers, row_idx, file_path, parent_base_addr, sheet_name, parent_size)
            else:
                row_modules = self._parse_addr_map_row_openpyxl_expanded(row_data, headers, row_idx, file_path, parent_base_addr, sheet_name, parent_size)

            if row_modules:
                modules.extend(row_modules)

        # 复制基模块的寄存器和子模块到数组实例
        for i, module in enumerate(modules):
            base_name = getattr(module, 'base_module_name', None)
            if base_name and base_name in self.all_modules:
                base_module = self.all_modules[base_name]

                # 检查是否需要复制寄存器或子模块
                needs_regs = base_module.registers and not module.registers
                needs_submods = base_module.submodules and not module.submodules

                if needs_regs or needs_submods:
                    # 复制寄存器
                    copied_regs = []
                    if needs_regs:
                        for reg in base_module.registers:
                            copied_fields = [
                                RegisterField(
                                    name=f.name, msb=f.msb, lsb=f.lsb,
                                    access=f.access, reset_value=f.reset_value,
                                    description=f.description
                                ) for f in reg.fields
                            ]
                            copied_regs.append(Register(
                                name=reg.name, offset=reg.offset, width=reg.width,
                                fields=copied_fields, description=reg.description
                            ))

                    # 复制子模块（递归处理）
                    copied_submods = list(module.submodules)  # 保留原有的
                    if needs_submods:
                        for sub in base_module.submodules:
                            copied_submod = self._copy_module_recursive(sub)
                            copied_submods.append(copied_submod)

                    modules[i] = Module(
                        name=module.name,
                        start_addr=module.start_addr,
                        end_addr=module.end_addr,
                        size=module.size,
                        registers=copied_regs if needs_regs else module.registers,
                        submodules=copied_submods,
                        is_array=module.is_array,
                        array_count=module.array_count,
                        source_file=module.source_file,
                        description=module.description,
                        is_array_instance=module.is_array_instance,
                        base_module_name=base_name
                    )
                    self.all_modules[module.name] = modules[i]

        # 检查地址/大小一致性问题
        self._validate_module_address_ranges(modules, sheet_name)

        return modules

    def _validate_module_address_ranges(self, modules: List[Module], context: str = ""):
        """检查模块的地址范围是否与其子模块一致

        检测问题:
        1. 子模块的总大小超过父模块的大小
        2. 子模块的地址范围超出父模块的地址范围
        """
        for module in modules:
            if not module.submodules:
                continue

            # 计算子模块占用的范围
            min_child_start = min(sub.start_addr for sub in module.submodules)
            max_child_end = max(sub.end_addr for sub in module.submodules)
            child_total_size = max_child_end - min_child_start + 1

            # 跳过相对地址的子模块（子模块起始地址小于父模块，说明是相对地址）
            if min_child_start < module.start_addr:
                # 子模块使用相对地址，只检查大小是否超出
                if child_total_size > module.size:
                    self.warnings.append(
                        f"地址/大小不一致警告 [{context}]: 模块'{module.name}'的大小为0x{module.size:X}({module.size//1024}K), "
                        f"但子模块占用的总大小为{child_total_size//1024}K。"
                        f"建议调整'{module.name}'的大小。"
                    )
            else:
                # 子模块使用绝对地址，检查是否超出父模块范围
                if child_total_size > module.size:
                    self.warnings.append(
                        f"地址/大小不一致警告 [{context}]: 模块'{module.name}'的大小为0x{module.size:X}({module.size//1024}K), "
                        f"但子模块占用的总范围为0x{min_child_start:X}-0x{max_child_end:X}(共{child_total_size//1024}K)。"
                        f"建议调整'{module.name}'的大小或子模块的地址。"
                    )

                if max_child_end > module.end_addr:
                    self.warnings.append(
                        f"地址/大小不一致警告 [{context}]: 模块'{module.name}'的地址范围为0x{module.start_addr:X}-0x{module.end_addr:X}, "
                        f"但子模块结束地址为0x{max_child_end:X}, 超出了父模块的范围。"
                    )

            # 递归检查子模块
            self._validate_module_address_ranges(module.submodules, context)

    def _copy_module_recursive(self, module: Module, visited: set = None) -> Module:
        """递归复制模块及其子模块"""
        if visited is None:
            visited = set()

        # 防止循环引用导致无限递归
        if module.name in visited:
            # 返回一个简化版的模块，避免循环
            return Module(
                name=module.name,
                start_addr=module.start_addr,
                end_addr=module.end_addr,
                size=module.size,
                registers=[],
                submodules=[],
                is_array=module.is_array,
                array_count=module.array_count,
                source_file=module.source_file,
                description=module.description
            )
        visited.add(module.name)

        # 复制寄存器
        copied_regs = []
        for reg in module.registers:
            copied_fields = [
                RegisterField(
                    name=f.name, msb=f.msb, lsb=f.lsb,
                    access=f.access, reset_value=f.reset_value,
                    description=f.description
                ) for f in reg.fields
            ]
            copied_regs.append(Register(
                name=reg.name, offset=reg.offset, width=reg.width,
                fields=copied_fields, description=reg.description
            ))

        # 递归复制子模块
        copied_submods = []
        for sub in module.submodules:
            copied_submods.append(self._copy_module_recursive(sub, visited.copy()))

        return Module(
            name=module.name,
            start_addr=module.start_addr,
            end_addr=module.end_addr,
            size=module.size,
            registers=copied_regs,
            submodules=copied_submods,
            is_array=module.is_array,
            array_count=module.array_count,
            source_file=module.source_file,
            description=module.description
        )

    def _parse_addr_map_row_xlrd(self, sheet, headers, row_idx, file_path) -> Optional[Module]:
        """解析addr_map的一行（xlrd）"""
        mod_name = str(sheet.cell_value(row_idx, headers.get('module_name', 0))).strip()
        start_str = str(sheet.cell_value(row_idx, headers.get('start_addr', 1))).strip()
        end_str = str(sheet.cell_value(row_idx, headers.get('end_addr', 2))).strip() if 'end_addr' in headers else ""
        size_str = str(sheet.cell_value(row_idx, headers.get('size', 3))).strip() if 'size' in headers else ""

        if not mod_name:
            return None

        # 解析模块数组（如 CPD*N(N=4)）
        is_array, array_count, base_name = self._parse_module_array(mod_name)

        start_addr = self._parse_address(start_str)
        end_addr = self._parse_address(end_str) if end_str else 0
        size = self._parse_size(size_str) if size_str else (end_addr - start_addr + 1)

        # 查找该模块的register表
        registers = self._find_registers_for_module(base_name)

        # 创建模块
        module = Module(
            name=mod_name,
            start_addr=start_addr,
            end_addr=end_addr if end_addr else start_addr + size - 1,
            size=size,
            registers=registers,
            is_array=is_array,
            array_count=array_count,
            source_file=file_path
        )

        self.all_modules[base_name] = module

        # 查找子模块的addr_map
        if base_name in self.addr_map_tables:
            sub_file, sub_sheet, sub_data = self.addr_map_tables[base_name]
            module.submodules = self._build_hierarchy_from_addr_map((sub_file, sub_sheet, sub_data))

        return module

    def _parse_addr_map_row_openpyxl(self, row, headers, row_idx, file_path) -> Optional[Module]:
        """解析addr_map的一行（openpyxl）"""
        mod_name = str(row[headers.get('module_name', 0)]).strip() if row[headers.get('module_name', 0)] else ""
        start_str = str(row[headers.get('start_addr', 1)]).strip() if row[headers.get('start_addr', 1)] else ""
        end_str = str(row[headers.get('end_addr', 2)]).strip() if 'end_addr' in headers and row[headers.get('end_addr', 2)] else ""
        size_str = str(row[headers.get('size', 3)]).strip() if 'size' in headers and row[headers.get('size', 3)] else ""

        if not mod_name:
            return None

        is_array, array_count, base_name = self._parse_module_array(mod_name)

        start_addr = self._parse_address(start_str)
        end_addr = self._parse_address(end_str) if end_str else 0
        size = self._parse_size(size_str) if size_str else (end_addr - start_addr + 1)

        registers = self._find_registers_for_module(base_name)

        module = Module(
            name=mod_name,
            start_addr=start_addr,
            end_addr=end_addr if end_addr else start_addr + size - 1,
            size=size,
            registers=registers,
            is_array=is_array,
            array_count=array_count,
            source_file=file_path
        )

        self.all_modules[base_name] = module

        if base_name in self.addr_map_tables:
            sub_file, sub_sheet, sub_data = self.addr_map_tables[base_name]
            module.submodules = self._build_hierarchy_from_addr_map((sub_file, sub_sheet, sub_data))

        return module

    def _validate_module_name(self, mod_name: str, row_idx: int, file_path: str) -> bool:
        """验证模块名是否合法"""
        if not mod_name:
            return False
        # 检查中文字符
        import re
        if re.search(r'[\u4e00-\u9fff]', mod_name):
            self.warnings.append(f"警告: {file_path} 第{row_idx}行模块名'{mod_name}'包含中文字符")
        # 检查中文括号
        if '（' in mod_name or '）' in mod_name:
            self.warnings.append(f"警告: {file_path} 第{row_idx}行模块名'{mod_name}'包含中文括号，已自动转换")
        return True

    def _validate_address(self, addr_str: str, addr_name: str, row_idx: int, file_path: str) -> bool:
        """验证地址格式"""
        if not addr_str:
            self.errors.append(f"错误: {file_path} 第{row_idx}行 {addr_name} 为空")
            return False
        addr_str = str(addr_str).strip()
        # 检查是否包含中文字符
        import re
        if re.search(r'[\u4e00-\u9fff]', addr_str):
            self.errors.append(f"错误: {file_path} 第{row_idx}行 {addr_name} '{addr_str}' 包含中文字符")
            return False
        return True

    def _parse_addr_map_row_xlrd_expanded(self, sheet, headers, row_idx, file_path, parent_base_addr: int = 0, sheet_name: str = "", parent_size: int = 0) -> List[Module]:
        """解析addr_map的一行（xlrd），展开模块数组"""
        # 读取并去除所有空格，转换中文括号
        mod_name = str(sheet.cell_value(row_idx, headers.get('module_name', 0))).replace(' ', '').replace('（', '(').replace('）', ')')
        start_str = str(sheet.cell_value(row_idx, headers.get('start_addr', 1))).replace(' ', '')
        end_str = str(sheet.cell_value(row_idx, headers.get('end_addr', 2))).replace(' ', '') if 'end_addr' in headers else ""
        size_str = str(sheet.cell_value(row_idx, headers.get('size', 3))).replace(' ', '') if 'size' in headers else ""

        if not mod_name:
            return []

        context = f"{file_path} 第{row_idx}行"
        if not sheet_name:
            sheet_name = sheet.name

        # 验证MODULE_NAME格式
        if not self.validator.validate_module_name(mod_name, context):
            return []

        # 验证MODULE_NAME不能和页签同名
        if not self.validator.validate_module_not_same_as_sheet(mod_name, sheet_name, context):
            return []

        # 解析模块数组（如 CPD*N(N=4)）
        is_array, array_count, base_name = self._parse_module_array(mod_name)

        # 验证地址格式
        valid, start_addr = self.validator.validate_address_format(start_str, 'start_addr', context)
        if not valid:
            return []
        start_addr += parent_base_addr

        # 验证end_addr格式
        end_addr = 0
        if end_str:
            valid, end_addr = self.validator.validate_address_format(end_str, 'end_addr', context, check_alignment=False)
            if not valid:
                return []
            end_addr += parent_base_addr

        # 验证size格式
        if size_str:
            valid, size = self.validator.validate_size_format(size_str, context)
            if not valid:
                return []
        else:
            size = end_addr - start_addr + 1 if end_addr else 0x1000

        # 验证地址范围 (end - start + 1 == size)
        if end_addr > 0 and not self.validator.validate_address_range(start_addr, end_addr, size, context):
            return []

        # 查找该模块的register表
        registers = self._find_registers_for_module(base_name)

        # 如果是数组但找不到寄存器表，给出警告
        if is_array and array_count > 1 and not registers:
            self.warnings.append(f"警告: {file_path} 模块'{base_name}'是数组但找不到寄存器定义")

        modules = []

        # 检查是否是数组实例（名称以数字结尾，如 C2C0, PE1）
        # 这种情况也需要创建基础模块+实例的模式
        is_named_instance = is_array and array_count == 1 and mod_name != base_name

        if (is_array and array_count > 1) or is_named_instance:
            # 首先创建/确保基础模块存在（只创建一份实现）
            # 基模块的地址从0开始，子模块的地址相对于基模块
            # 根据parent_size和address_range判断size是per-instance还是total
            address_range = end_addr - start_addr + 1 if end_addr > start_addr else size
            if parent_size > 0 and address_range * array_count > parent_size:
                # size是total，计算per-instance size (使用address_range作为total size)
                instance_size = address_range // array_count if array_count > 0 else address_range
            else:
                # size是per-instance
                instance_size = size if size else 0x1000
            if base_name not in self.all_modules:
                base_module = Module(
                    name=base_name,
                    start_addr=0,  # 基模块地址为0，子模块地址相对于基模块
                    end_addr=instance_size - 1,
                    size=instance_size,
                    registers=list(registers) if registers else [],  # Copy to avoid reference sharing
                    is_array=True,
                    array_count=array_count,
                    source_file=file_path
                )
                # 查找子模块的addr_map（用于基础模块）
                if base_name in self.addr_map_tables:
                    sub_file, sub_sheet, sub_data = self.addr_map_tables[base_name]
                    base_module.submodules = self._build_hierarchy_from_addr_map((sub_file, sub_sheet, sub_data), 0, 0)
                self.all_modules[base_name] = base_module

            # 展开模块数组，创建多个实例（不复制寄存器，只记录引用）
            # 对于命名实例（如C2C0），使用原始名称作为实例名
            if is_named_instance:
                # 单个命名实例，直接使用原始名称
                instance_name = mod_name
                instance_start = start_addr
                instance_end = end_addr if end_addr else (start_addr + instance_size - 1)

                # 从基础模块获取子模块（如果基础模块已存在）
                base_submodules = self.all_modules[base_name].submodules if base_name in self.all_modules else []

                instance = Module(
                    name=instance_name,
                    start_addr=instance_start,
                    end_addr=instance_end,
                    size=instance_size,
                    registers=[],  # 实例不直接包含寄存器，引用基础模块
                    submodules=base_submodules,  # 复制子模块以维持层次结构
                    is_array=False,
                    array_count=1,
                    source_file=file_path,
                    is_array_instance=True,
                    base_module_name=base_name
                )
                self.all_modules[instance_name] = instance
                modules.append(instance)
            else:
                # 多个实例的情况（如 *N(N=4)）
                # 从基础模块获取子模块（如果基础模块已存在）
                base_submodules = self.all_modules[base_name].submodules if base_name in self.all_modules else []
                for i in range(array_count):
                    instance_name = f"{base_name}_{i}"
                    instance_start = start_addr + i * instance_size
                    instance_end = instance_start + instance_size - 1

                    # 创建实例（不复制寄存器，引用基础模块）
                    instance = Module(
                        name=instance_name,
                        start_addr=instance_start,
                        end_addr=instance_end,
                        size=instance_size,
                        registers=[],  # 实例不直接包含寄存器，引用基础模块
                        submodules=base_submodules,  # 复制子模块以维持层次结构
                        is_array=False,
                        array_count=1,
                        source_file=file_path,
                        is_array_instance=True,
                        base_module_name=base_name
                    )

                    self.all_modules[instance_name] = instance
                    modules.append(instance)
        else:
            # 单个模块
            module = Module(
                name=base_name,
                start_addr=start_addr,
                end_addr=end_addr if end_addr else start_addr + size - 1,
                size=size,
                registers=registers,
                is_array=is_array,
                array_count=array_count,
                source_file=file_path
            )

            self.all_modules[base_name] = module

            # 查找子模块的addr_map
            if base_name in self.addr_map_tables:
                sub_file, sub_sheet, sub_data = self.addr_map_tables[base_name]
                module.submodules = self._build_hierarchy_from_addr_map((sub_file, sub_sheet, sub_data), start_addr)

            modules.append(module)

        return modules

    def _parse_addr_map_row_openpyxl_expanded(self, row, headers, row_idx, file_path, parent_base_addr: int = 0, sheet_name: str = "", parent_size: int = 0) -> List[Module]:
        """解析addr_map的一行（openpyxl），展开模块数组"""
        # 读取并去除所有空格，转换中文括号
        mod_name = str(row[headers.get('module_name', 0)]).replace(' ', '').replace('（', '(').replace('）', ')') if row[headers.get('module_name', 0)] else ""
        start_str = str(row[headers.get('start_addr', 1)]).replace(' ', '') if row[headers.get('start_addr', 1)] else ""
        end_str = str(row[headers.get('end_addr', 2)]).replace(' ', '') if 'end_addr' in headers and row[headers.get('end_addr', 2)] else ""
        size_str = str(row[headers.get('size', 3)]).replace(' ', '') if 'size' in headers and row[headers.get('size', 3)] else ""

        if not mod_name:
            return []

        context = f"{file_path} 第{row_idx}行"

        # 验证MODULE_NAME格式
        if not self.validator.validate_module_name(mod_name, context):
            return []

        # 验证MODULE_NAME不能和页签同名
        if sheet_name and not self.validator.validate_module_not_same_as_sheet(mod_name, sheet_name, context):
            return []

        is_array, array_count, base_name = self._parse_module_array(mod_name)

        # 验证地址格式
        valid, start_addr = self.validator.validate_address_format(start_str, 'start_addr', context)
        if not valid:
            return []
        start_addr += parent_base_addr

        # 验证end_addr格式
        end_addr = 0
        if end_str:
            valid, end_addr = self.validator.validate_address_format(end_str, 'end_addr', context, check_alignment=False)
            if not valid:
                return []
            end_addr += parent_base_addr

        # 验证size格式
        if size_str:
            valid, size = self.validator.validate_size_format(size_str, context)
            if not valid:
                return []
        else:
            size = end_addr - start_addr + 1 if end_addr else 0x1000

        # 验证地址范围 (end - start + 1 == size)
        if end_addr > 0 and not self.validator.validate_address_range(start_addr, end_addr, size, context):
            return []

        registers = self._find_registers_for_module(base_name)

        modules = []

        # 检查是否是数组实例（名称以数字结尾，如 C2C0, PE1）
        # 这种情况也需要创建基础模块+实例的模式
        is_named_instance = is_array and array_count == 1 and mod_name != base_name

        if (is_array and array_count > 1) or is_named_instance:
            # 首先创建/确保基础模块存在（只创建一份实现）
            # 基模块的地址从0开始，子模块的地址相对于基模块
            # 根据parent_size和address_range判断size是per-instance还是total
            address_range = end_addr - start_addr + 1 if end_addr > start_addr else size
            if parent_size > 0 and address_range * array_count > parent_size:
                # size是total，计算per-instance size (使用address_range作为total size)
                instance_size = address_range // array_count if array_count > 0 else address_range
            else:
                # size是per-instance
                instance_size = size if size else 0x1000
            if base_name not in self.all_modules:
                base_module = Module(
                    name=base_name,
                    start_addr=0,  # 基模块地址为0，子模块地址相对于基模块
                    end_addr=instance_size - 1,
                    size=instance_size,
                    registers=list(registers) if registers else [],  # Copy to avoid reference sharing
                    is_array=True,
                    array_count=array_count,
                    source_file=file_path
                )
                # 查找子模块的addr_map（用于基础模块）
                if base_name in self.addr_map_tables:
                    sub_file, sub_sheet, sub_data = self.addr_map_tables[base_name]
                    base_module.submodules = self._build_hierarchy_from_addr_map((sub_file, sub_sheet, sub_data), 0, 0)
                self.all_modules[base_name] = base_module

            # 展开模块数组，创建多个实例（不复制寄存器，只记录引用）
            # 对于命名实例（如C2C0），使用原始名称作为实例名
            if is_named_instance:
                # 单个命名实例，直接使用原始名称
                instance_name = mod_name
                instance_start = start_addr
                instance_end = end_addr if end_addr else (start_addr + size - 1)
                instance_size = size

                # 从基础模块获取子模块（如果基础模块已存在）
                base_submodules = self.all_modules[base_name].submodules if base_name in self.all_modules else []

                instance = Module(
                    name=instance_name,
                    start_addr=instance_start,
                    end_addr=instance_end,
                    size=instance_size,
                    registers=[],  # 实例不直接包含寄存器，引用基础模块
                    submodules=base_submodules,  # 复制子模块以维持层次结构
                    is_array=False,
                    array_count=1,
                    source_file=file_path,
                    is_array_instance=True,
                    base_module_name=base_name
                )
                self.all_modules[instance_name] = instance
                modules.append(instance)
            else:
                # 多个实例的情况（如 *N(N=4)）
                # 从基础模块获取子模块（如果基础模块已存在）
                base_submodules = self.all_modules[base_name].submodules if base_name in self.all_modules else []
                for i in range(array_count):
                    instance_name = f"{base_name}_{i}"
                    instance_start = start_addr + i * instance_size
                    instance_end = instance_start + instance_size - 1

                    # 创建实例（不复制寄存器，引用基础模块）
                    instance = Module(
                        name=instance_name,
                        start_addr=instance_start,
                        end_addr=instance_end,
                        size=instance_size,
                        registers=[],  # 实例不直接包含寄存器，引用基础模块
                        submodules=base_submodules,  # 复制子模块以维持层次结构
                        is_array=False,
                        array_count=1,
                        source_file=file_path,
                        is_array_instance=True,
                        base_module_name=base_name
                    )

                    self.all_modules[instance_name] = instance
                    modules.append(instance)
        else:
            module = Module(
                name=base_name,
                start_addr=start_addr,
                end_addr=end_addr if end_addr else start_addr + size - 1,
                size=size,
                registers=registers,
                is_array=is_array,
                array_count=array_count,
                source_file=file_path
            )

            self.all_modules[base_name] = module

            if base_name in self.addr_map_tables:
                sub_file, sub_sheet, sub_data = self.addr_map_tables[base_name]
                module.submodules = self._build_hierarchy_from_addr_map((sub_file, sub_sheet, sub_data), start_addr)

            modules.append(module)

        return modules

    def _copy_registers_with_offset(self, registers: List[Register], offset: int) -> List[Register]:
        """复制寄存器列表并更新地址偏移"""
        import copy
        new_regs = []
        for reg in registers:
            new_reg = copy.copy(reg)
            new_reg.offset = reg.offset + offset
            new_reg.fields = [copy.copy(f) for f in reg.fields]
            new_regs.append(new_reg)
        return new_regs

    def _parse_module_array(self, mod_name: str) -> Tuple[bool, int, str]:
        """解析模块数组，如 CPD*N(N=4) -> (True, 4, 'CPD') 或 PEC0 -> (True, 1, 'PEC')"""
        # 标准化中文括号为英文括号
        normalized_name = mod_name.replace('（', '(').replace('）', ')')

        # 匹配 *N(N=num) 或 *N(N=num) 格式（支持中文括号）
        match = re.match(r'^(.*?)\*\s*N\s*\(\s*N\s*=\s*(\d+)\s*\)$', normalized_name, re.IGNORECASE)
        if match:
            base_name = match.group(1).strip()
            # 去掉末尾的下划线
            if base_name.endswith('_'):
                base_name = base_name[:-1]
            return True, int(match.group(2)), base_name

        # 匹配 Name0/Name1 格式 - 数字后缀视为数组实例
        # 例如 PEC0 -> 基模块 PEC, PEC1 -> 基模块 PEC
        match = re.match(r'^([A-Za-z_][A-Za-z0-9_]*?)(\d+)$', mod_name)
        if match:
            base_name = match.group(1)
            # 如果base_name以下划线结尾，去掉它
            if base_name.endswith('_'):
                base_name = base_name[:-1]
            # 视为单元素数组实例（数组大小为1，但标记为数组类型）
            return True, 1, base_name

        return False, 1, mod_name

    def _parse_address(self, addr_str: str) -> int:
        """解析地址字符串"""
        if not addr_str:
            return 0
        addr_str = str(addr_str).strip().replace('_', '').replace(' ', '')

        if addr_str.lower().startswith('0x'):
            try:
                return int(addr_str, 16)
            except:
                return 0
        else:
            try:
                return int(addr_str)
            except:
                return 0

    def _parse_size(self, size_str: str) -> int:
        """解析大小字符串，如 8K, 4M, 8KB, 4MB"""
        if not size_str:
            return 0
        size_str = str(size_str).strip().upper().replace(' ', '').replace('B', '')

        if size_str.endswith('K'):
            try:
                return int(size_str[:-1]) * 1024
            except:
                return 0
        elif size_str.endswith('M'):
            try:
                return int(size_str[:-1]) * 1024 * 1024
            except:
                return 0
        elif size_str.endswith('G'):
            try:
                return int(size_str[:-1]) * 1024 * 1024 * 1024
            except:
                return 0
        else:
            try:
                return int(size_str)
            except:
                return 0

    def _find_registers_for_module(self, module_name: str) -> List[Register]:
        """查找模块的register表

        支持ABC_0/ABC_1或ABC0/ABC1格式，会尝试寻找ABC为名的module
        """
        import re

        # 首先尝试完整匹配
        if module_name in self.register_tables:
            file_path, sheet_name, data = self.register_tables[module_name]
            return self._parse_register_sheet(data, module_name)

        # 尝试ABC_0/ABC_1 -> ABC匹配
        # 匹配模式: NAME_0, NAME_1, NAME0, NAME1
        patterns = [
            r'^(\w+)_\d+$',      # ABC_0, ABC_1
            r'^(\w+?)(\d+)$',    # ABC0, ABC1
        ]

        for pattern in patterns:
            match = re.match(pattern, module_name)
            if match:
                base_name = match.group(1)
                if base_name in self.register_tables:
                    file_path, sheet_name, data = self.register_tables[base_name]
                    return self._parse_register_sheet(data, base_name)

        return []

    def _parse_register_sheet(self, data: Dict, module_name: str, base_addr: int = 0) -> List[Register]:
        """解析register表格，支持寄存器数组展开"""
        registers = []
        sheet = data.get('xlrd') or data.get('openpyxl')
        headers = data['headers']

        # 临时存储：原始寄存器名 -> (base_offset, width, description, fields)
        reg_groups = []
        current_group = None

        def parse_row(row_data, is_xlrd=True):
            """解析一行数据 - 去除所有空格"""
            if is_xlrd:
                offset_str = str(sheet.cell_value(row_data, headers.get('offset_address', 0))).replace(' ', '')
                reg_name = str(sheet.cell_value(row_data, headers.get('reg_name', 1))).replace(' ', '')
                width_str = str(sheet.cell_value(row_data, headers.get('width', 2))).replace(' ', '')
                field_name = str(sheet.cell_value(row_data, headers.get('field_name', 4))).replace(' ', '')
                access = str(sheet.cell_value(row_data, headers.get('access', 5))).replace(' ', '')
                reset_str = str(sheet.cell_value(row_data, headers.get('reset_value', 6))).replace(' ', '')
                desc = str(sheet.cell_value(row_data, headers.get('description', 7))).strip() if 'description' in headers else ""
                bits_str = str(sheet.cell_value(row_data, headers['bits'])).replace(' ', '') if 'bits' in headers else ""
                msb_val = sheet.cell_value(row_data, headers['msb']) if 'msb' in headers else None
                lsb_val = sheet.cell_value(row_data, headers['lsb']) if 'lsb' in headers else None
            else:
                # Helper to safely get cell value
                def safe_get(idx):
                    return row_data[idx] if idx < len(row_data) else None

                offset_str = str(safe_get(headers.get('offset_address', 0))).replace(' ', '') if safe_get(headers.get('offset_address', 0)) else ""
                reg_name = str(safe_get(headers.get('reg_name', 1))).replace(' ', '') if safe_get(headers.get('reg_name', 1)) else ""
                width_str = str(safe_get(headers.get('width', 2))).replace(' ', '') if safe_get(headers.get('width', 2)) else ""
                field_name = str(safe_get(headers.get('field_name', 4))).replace(' ', '') if safe_get(headers.get('field_name', 4)) else ""
                access = str(safe_get(headers.get('access', 5))).replace(' ', '') if safe_get(headers.get('access', 5)) else ""
                reset_str = str(safe_get(headers.get('reset_value', 6))).replace(' ', '') if safe_get(headers.get('reset_value', 6)) else ""
                desc = str(safe_get(headers.get('description', 7))).strip() if 'description' in headers and safe_get(headers.get('description', 7)) else ""
                bits_str = str(safe_get(headers['bits'])).replace(' ', '') if 'bits' in headers and safe_get(headers['bits']) else ""
                msb_val = safe_get(headers['msb']) if 'msb' in headers else None
                lsb_val = safe_get(headers['lsb']) if 'lsb' in headers else None

            return offset_str, reg_name, width_str, field_name, access, reset_str, desc, bits_str, msb_val, lsb_val

        def safe_int(val, default=0):
            """安全地转换为整数，支持十六进制"""
            if val is None:
                return default
            try:
                if isinstance(val, str):
                    val = val.strip().replace("'h", "0x").replace("'b", "0b").replace("'d", "")
                    if val.startswith('0x') or val.startswith('0X'):
                        return int(val, 16)
                    elif val.startswith('0b') or val.startswith('0B'):
                        return int(val, 2)
                    elif val:
                        return int(float(val))
                    return default
                return int(val)
            except:
                return default

        def calc_msb_lsb(bits_str, msb_val, lsb_val):
            """计算msb/lsb"""
            if bits_str:
                return self._parse_bits(bits_str)
            elif msb_val is not None and lsb_val is not None:
                try:
                    return safe_int(msb_val), safe_int(lsb_val)
                except:
                    return 0, 0
            return 0, 0

        # 第一遍：收集所有寄存器和字段定义
        # 使用字典来跟踪已经创建的寄存器组，同名寄存器（去除数组标记后）共享同一组
        group_map = {}  # (base_name, offset) -> group
        # 位宽继承：整张表格的位宽必须一致，第一个寄存器的位宽会被后续空位宽的寄存器继承
        inherited_width = 32  # 默认32位

        if 'xlrd' in data:
            for row_idx in range(1, sheet.nrows):
                offset_str, reg_name, width_str, field_name, access, reset_str, desc, bits_str, msb_val, lsb_val = parse_row(row_idx, True)

                if reg_name:
                    context = f"{module_name} 第{row_idx}行"
                    offset = self._parse_address(offset_str) if offset_str else 0

                    # 验证width，支持位宽继承（整张表格位宽一致）
                    if width_str:
                        valid, width = self.validator.validate_width(width_str, context)
                        if not valid:
                            self.errors.extend(self.validator.errors)
                            self.validator.clear()
                        # 更新继承位宽
                        inherited_width = width
                    else:
                        # 继承之前的位宽
                        width = inherited_width

                    # 验证64位地址对齐
                    if width == 64:
                        if not self.validator.validate_64bit_address_alignment(offset, width, context):
                            self.errors.extend(self.validator.errors)
                            self.validator.clear()

                    # 解析数组
                    base_name, array_count = self._parse_reg_array(reg_name)
                    # 使用 (base_name, offset) 作为唯一键
                    group_key = (base_name, offset)
                    if group_key not in group_map:
                        current_group = {
                            'name': base_name,
                            'offset': offset,
                            'width': width,
                            'description': desc,
                            'array_count': array_count,
                            'fields': []
                        }
                        group_map[group_key] = current_group
                        reg_groups.append(current_group)
                    else:
                        current_group = group_map[group_key]

                if current_group and field_name:
                    msb, lsb = calc_msb_lsb(bits_str, msb_val, lsb_val)

                    # 验证access类型
                    context = f"{module_name} 字段'{field_name}'"
                    valid, access_type = self.validator.validate_access_type(access, context)
                    if not valid:
                        self.errors.extend(self.validator.errors)
                        self.validator.clear()

                    field = {
                        'name': field_name,
                        'msb': msb,
                        'lsb': lsb,
                        'access': access_type,
                        'reset_value': reset_str,
                        'description': desc
                    }
                    current_group['fields'].append(field)
        else:  # openpyxl
            row_idx = 1  # Track row index for error messages
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_idx += 1
                offset_str, reg_name, width_str, field_name, access, reset_str, desc, bits_str, msb_val, lsb_val = parse_row(row, False)

                if reg_name:
                    context = f"{module_name} 第{row_idx}行"
                    offset = self._parse_address(offset_str) if offset_str else 0

                    # 验证width，支持位宽继承（整张表格位宽一致）
                    if width_str:
                        valid, width = self.validator.validate_width(width_str, context)
                        if not valid:
                            self.errors.extend(self.validator.errors)
                            self.validator.clear()
                        # 更新继承位宽
                        inherited_width = width
                    else:
                        # 继承之前的位宽
                        width = inherited_width

                    # 验证64位地址对齐
                    if width == 64:
                        if not self.validator.validate_64bit_address_alignment(offset, width, context):
                            self.errors.extend(self.validator.errors)
                            self.validator.clear()

                    base_name, array_count = self._parse_reg_array(reg_name)
                    # 使用 (base_name, offset) 作为唯一键
                    group_key = (base_name, offset)
                    if group_key not in group_map:
                        current_group = {
                            'name': base_name,
                            'offset': offset,
                            'width': width,
                            'description': desc,
                            'array_count': array_count,
                            'fields': []
                        }
                        group_map[group_key] = current_group
                        reg_groups.append(current_group)
                    else:
                        current_group = group_map[group_key]

                if current_group and field_name:
                    msb, lsb = calc_msb_lsb(bits_str, msb_val, lsb_val)

                    # 验证access类型
                    context = f"{module_name} 字段'{field_name}'"
                    valid, access_type = self.validator.validate_access_type(access, context)
                    if not valid:
                        self.errors.extend(self.validator.errors)
                        self.validator.clear()

                    field = {
                        'name': field_name,
                        'msb': msb,
                        'lsb': lsb,
                        'access': access_type,
                        'reset_value': reset_str,
                        'description': desc
                    }
                    current_group['fields'].append(field)

        # 验证：检查同一offset的字段是否连续
        for group in reg_groups:
            fields = group['fields']
            if len(fields) > 1:
                # 按lsb排序
                sorted_fields = sorted(fields, key=lambda f: f['lsb'])
                regwidth = group['width']

                # 检查是否有重叠
                for i in range(len(sorted_fields) - 1):
                    curr = sorted_fields[i]
                    next_f = sorted_fields[i + 1]
                    if curr['msb'] >= next_f['lsb']:
                        self.errors.append(
                            f"错误: 寄存器 '{group['name']}' @0x{group['offset']:X} "
                            f"字段 '{curr['name']}'[{curr['msb']}:{curr['lsb']}] 和 "
                            f"'{next_f['name']}'[{next_f['msb']}:{next_f['lsb']}] 有重叠"
                        )

                # 检查是否填满整个寄存器宽度（字段不连续只警告，不报错）
                # 从lsb=0开始检查连续性
                expected_lsb = 0
                gaps = []
                for f in sorted_fields:
                    if f['lsb'] > expected_lsb:
                        gaps.append(f"[{f['lsb']-1}:{expected_lsb}]")
                    expected_lsb = f['msb'] + 1

                # 检查最高位到regwidth-1
                if expected_lsb < regwidth:
                    gaps.append(f"[{regwidth-1}:{expected_lsb}]")

                if gaps:
                    # 字段不连续只警告，不阻止生成（可能有保留位）
                    self.warnings.append(
                        f"寄存器 '{group['name']}' @0x{group['offset']:X} "
                        f"字段未填满整个寄存器宽度，缺失位域（可能是保留位）: {', '.join(gaps)}"
                    )

        # 只有字段重叠等严重错误才返回空列表
        if self.errors:
            return []

        # 第二遍：展开数组并创建寄存器对象
        for group in reg_groups:
            expanded = self._expand_reg_array(group['name'], group['offset'], group['width'], group['array_count'])
            for reg_name, reg_offset in expanded:
                reg = Register(
                    name=reg_name,
                    offset=reg_offset,
                    width=group['width'],
                    description=group['description']
                )
                # 复制字段到每个展开的寄存器
                for field_def in group['fields']:
                    field = RegisterField(
                        name=field_def['name'],
                        msb=field_def['msb'],
                        lsb=field_def['lsb'],
                        access=field_def['access'],
                        reset_value=field_def['reset_value'],
                        description=field_def['description']
                    )
                    reg.fields.append(field)
                registers.append(reg)

        return registers

    def _parse_bits(self, bits_str: str) -> Tuple[int, int]:
        """解析位范围，如 [31:24] 或 31:24"""
        bits_str = str(bits_str).strip().replace('[', '').replace(']', '')

        if ':' in bits_str:
            parts = bits_str.split(':')
            try:
                msb = int(parts[0])
                lsb = int(parts[1])
                return msb, lsb
            except:
                return 0, 0
        else:
            try:
                bit = int(bits_str)
                return bit, bit
            except:
                return 0, 0

    def _parse_reg_array(self, reg_name: str) -> Tuple[str, int]:
        """
        解析寄存器名称中的数组标记
        支持格式: REG_NAME*N(N=8) 或 REG_NAME*N(N=2)
        返回: (base_name, array_count)
        """
        import re
        if not reg_name:
            return reg_name, 1

        # 匹配 *N(N=数字) 格式
        pattern = r'^(.*?)\*N\s*\(\s*N\s*=\s*(\d+)\s*\)$'
        match = re.match(pattern, reg_name, re.IGNORECASE)
        if match:
            base_name = match.group(1).strip()
            array_count = int(match.group(2))
            return base_name, array_count

        # 尝试其他格式如 *N(N=2) 的变体
        pattern2 = r'^(.*?)\*\s*(\d+)$'
        match2 = re.match(pattern2, reg_name)
        if match2:
            base_name = match2.group(1).strip()
            array_count = int(match2.group(2))
            return base_name, array_count

        return reg_name, 1

    def _expand_reg_array(self, reg_name: str, base_offset: int, width: int, array_count: int) -> List[Tuple[str, int]]:
        """
        展开寄存器数组
        返回: [(reg_name_0, offset_0), (reg_name_1, offset_1), ...]
        """
        if array_count <= 1:
            return [(reg_name, base_offset)]

        results = []
        addr_step = 4 if width <= 32 else 8  # 32bit=+4, 64bit=+8

        for i in range(array_count):
            indexed_name = f"{reg_name}_{i}"  # 或 {reg_name}{i}
            offset = base_offset + (i * addr_step)
            results.append((indexed_name, offset))

        return results
